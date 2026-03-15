from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from .models import BeatmapEntry, CollectionInfo


EXPORT_HEADERS = [
    "Collection",
    "Mode",
    "Missing",
    "Name",
    "Artist",
    "Title",
    "BID",
    "SID",
    "Difficulty",
    "Mapper",
    "Background URL",
    "MD5",
]


def _sheet_name(base: str, used: set[str]) -> str:
    cleaned = (base or "Collection").replace("/", " ").replace("\\", " ").replace("*", " ").replace("?", " ")
    cleaned = cleaned.replace("[", "(").replace("]", ")").replace(":", " ").strip() or "Collection"
    cleaned = cleaned[:31]
    candidate = cleaned
    index = 1
    while candidate in used:
        suffix = f"_{index}"
        candidate = f"{cleaned[:31 - len(suffix)]}{suffix}"
        index += 1
    used.add(candidate)
    return candidate


def _write_rows(sheet, collection_name: str, items: list[BeatmapEntry]) -> None:
    sheet.append(EXPORT_HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for item in items:
        sheet.append(
            [
                collection_name,
                item.mode,
                "Yes" if item.missing else "No",
                item.name,
                item.artist,
                item.title,
                item.bid_text,
                item.sid_text,
                item.difficulty_name,
                item.mapper,
                item.background_url,
                item.md5,
            ]
        )

    widths = {
        "A": 24,
        "B": 12,
        "C": 10,
        "D": 38,
        "E": 22,
        "F": 24,
        "G": 12,
        "H": 12,
        "I": 24,
        "J": 18,
        "K": 56,
        "L": 36,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def export_collection(path: Path, collection: CollectionInfo, mode: str) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Collection"
    _write_rows(sheet, collection.name, collection.items_for_mode(mode))
    workbook.save(path)


def export_filtered(path: Path, collections: list[CollectionInfo], mode: str) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.append(["Collection", "Mode", "Visible Items", "Missing Items", "Last Modified"])
    for cell in summary[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    used_names = {"Summary"}
    for collection in collections:
        visible_items = collection.items_for_mode(mode)
        if not visible_items:
            continue

        summary.append(
            [
                collection.name,
                mode,
                len(visible_items),
                sum(1 for item in visible_items if item.missing),
                collection.last_modified_text,
            ]
        )

        sheet = workbook.create_sheet(_sheet_name(collection.name, used_names))
        _write_rows(sheet, collection.name, visible_items)

    for column, width in {"A": 28, "B": 12, "C": 14, "D": 14, "E": 22}.items():
        summary.column_dimensions[column].width = width

    workbook.save(path)


def export_current_view(
    path: Path,
    sheet_name: str,
    headers: list[str],
    rows: list[list[str]],
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _sheet_name(sheet_name or "Beatmaps", set())

    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        sheet.append(row)

    for column_cells in sheet.columns:
        letter = column_cells[0].column_letter
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[letter].width = min(max(max_length + 2, 10), 40)

    workbook.save(path)
